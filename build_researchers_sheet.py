#!/usr/bin/env python3
"""Build Autism_Researchers.xlsx — comprehensive researcher tracker
for the atlas. Three sheets:
  1. Researchers (main): 25 researchers, 14 columns
  2. Key Papers: top peer-reviewed publications per researcher
  3. Signal vs Noise: per-researcher quality assessment

Per CLAUDE.md mainstream-skeptical principle, the list balances mainstream
academics, functional medicine clinicians, contested-evidence researchers,
and parent-researchers. Signal density rating (1-10) reflects ratio of
primary-research output to commentary/engagement output.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUT = "Autism_Researchers.xlsx"

# Styles
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', start_color='1F4E79')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
BODY_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
THIN = Side(border_style='thin', color='888888')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Tier color coding
TIER_FILLS = {
    'Functional medicine / clinical': PatternFill('solid', start_color='E2EFDA'),  # light green
    'Contested / mainstream-skeptical': PatternFill('solid', start_color='FFF2CC'),  # light yellow
    'Mainstream academic': PatternFill('solid', start_color='D9E1F2'),  # light blue
    'Legal / policy': PatternFill('solid', start_color='FCE4D6'),  # light orange
    'Journalist / parent-advocate': PatternFill('solid', start_color='F4CCCC'),  # light pink
}

# === Researcher data ===
researchers = [
    # (name, credentials, affiliation, twitter, other_platform, parent_y_n, focus,
    #  contributions, atlas_connections, tier, signal, bias, reliability, notes)
    (1, "Richard E. Frye", "MD, PhD",
     "Rossignol Medical Center; Autism Discovery & Treatment Foundation",
     "@AutismEnergetic", "drfryemdphd.com",
     "N",
     "Cerebral folate deficiency; mitochondrial autism; FOLR1 autoantibodies; methylation; PANS",
     "300+ peer-reviewed publications. Lead investigator on multiple leucovorin RCTs. Frye/Slattery FOLR1 protocol. The atlas's calibration anchor scientist (INT-0001 leucovorin = 83.20 CSRS based on his work).",
     "INT-0001; HYP-0001; PHE-0001; SRC-000001; SRC-000011-26 (multiple)",
     "Functional medicine / clinical",
     9,
     "Functional medicine practitioner; criticized by mainstream as advocacy-leaning, but his peer-reviewed track record is solid. Owns clinic that uses interventions he researches (financial overlap acknowledged in disclosures).",
     "High",
     "STRONGEST signal in the list. Almost all his Twitter posts are research updates or interpretations of his own/others' published work. Subscribe priority #1."),

    (2, "Daniel Rossignol", "MD",
     "Rossignol Medical Center",
     "(handle uncertain; primarily speaks via clinic + podcasts)", "rossignolmedicalcenter.com",
     "N",
     "HBOT for autism; mitochondrial autism; functional medicine; methyl-B12 protocols",
     "55+ peer-reviewed papers. Long-term Frye collaborator. Founded Rossignol Medical Center.",
     "INT-0092 (HBOT); HYP-0001; SRC-000023 (and others)",
     "Functional medicine / clinical",
     7,
     "Same financial overlap concern as Frye. Less Twitter-active.",
     "High",
     "Substantive primary research output but less social media presence. Follow primarily via published papers."),

    (3, "Robert K. Naviaux", "MD, PhD",
     "UCSD Mitochondrial and Metabolic Disease Center",
     "(rare Twitter use; primary platform: lab website)", "naviauxlab.ucsd.edu",
     "N",
     "Cell danger response (CDR) theory; suramin antipurinergic clinical trial; mitochondrial autism mechanism; metabolomics",
     "Originated cell danger response framework. SAT1 trial: low-dose suramin in autism (2017 small open-label trial showed effect; replication pending). Foundational metabolomics in ASD.",
     "MEC-0010 (mito); HYP-0001 (CDR overlaps cerebral folate); future drug repurposing entries (suramin)",
     "Mainstream academic",
     8,
     "Academic, not advocacy-aligned. Suramin's contested-but-promising status reflects single small trial, not Naviaux's reputation issue.",
     "High",
     "Career academic. Primary research output is the signal. Subscribe to his lab page rather than Twitter."),

    (4, "Theoharis Theoharides", "MD, PhD",
     "Tufts (emeritus); Institute for Neuro-Immune Medicine",
     "(rare Twitter use; primary platform: mastcellmaster.com)", "mastcellmaster.com",
     "N",
     "Mast cell activation in autism; brain-mast-cell axis; flavonoid therapeutics (luteolin, quercetin)",
     "h-index 113. Foundational MCAS-autism researcher. Showed mast cell mediators cross BBB and affect neuroinflammation.",
     "MEC-0017 (mast cell activation); planned HYP for MCAS in Session 3.5C; INT-0029 (quercetin)",
     "Functional medicine / clinical",
     8,
     "Patents on luteolin formulation (NeuroProtek); financial interest disclosed.",
     "High",
     "Foundational researcher for MCAS-autism subset. Atlas Session 3.5C will ingest his framework as a new HYP."),

    (5, "Sabine Hazan", "MD",
     "ProgenaBiome; Microbiome Research Foundation",
     "@SabinehazanMD", "progenabiome.com",
     "N",
     "Gut microbiome in autism; Bifidobacterium infantis depletion; FMT for autism; MTT (microbiota transfer therapy)",
     "Active researcher on autism gut microbiome. Numerous Twitter/X posts citing primary microbiome work. Collaborates with Adams/Krajmalnik-Brown on MTT studies.",
     "MEC-0008 (gut-brain axis); HYP-0007; HYP-0056 (B. infantis); INT-0025 (probiotics); INT-0076 (FMT); INT-0077",
     "Functional medicine / clinical",
     7,
     "Owns ProgenaBiome (microbiome lab/clinic); financial interest in microbiome interventions.",
     "Medium-High",
     "Strong primary research signal on microbiome work; some posts are commercial-adjacent. Filter by whether post links to a published paper or just opinion."),

    (6, "James B. Adams", "PhD",
     "Arizona State University; Autism Nutrition Research Center",
     "(less Twitter-active; primary platform: ASU lab + autismnrc.org)", "autismnrc.org",
     "Y",
     "Microbiota Transfer Therapy (MTT); ASU autism nutrition program; B-vitamins, carnitine, gut-brain",
     "150+ peer-reviewed papers. President's Professor at ASU. Adult autistic daughter is original research catalyst. Ran MTT studies showing 2-year sustained autism symptom reduction (~50%) post-FMT.",
     "INT-0076 (FMT); HYP-0007; HYP-0056; broad nutritional intervention coverage",
     "Functional medicine / clinical",
     9,
     "Academic with lived parental motivation. Long publication track record. No major financial conflicts.",
     "Very High",
     "Mainstream-published primary-research scientist with parental motivation. Highest reliability tier. Subscribe via ASU page."),

    (7, "Christopher Exley", "PhD",
     "Independent (formerly Keele University; effectively forced out 2021)",
     "(Substack primary platform; X presence limited)", "drchristopherexley.substack.com",
     "N",
     "Aluminum biochemistry; aluminum in autism brain tissue; aluminum adjuvants",
     "Foundational researcher on aluminum brain accumulation. Mochizuki-Exley 2018 showed elevated aluminum in autism brain tissue. Forced out of Keele after refusing to drop autism-aluminum research.",
     "HYP-0067 (aluminum adjuvant); MEC-0032; SRC-001431 (Tomljenovic-Shaw is collaborator-adjacent)",
     "Contested / mainstream-skeptical",
     8,
     "Heavily criticized by mainstream; primary research findings (high Al in autism brain) replicated. Career destroyed by maintaining position. Less productive output now due to losing institutional support.",
     "Medium-High",
     "Career-destroyed academic with substantive primary findings. Substack output is mix of advocacy + science updates."),

    (8, "James Lyons-Weiler", "PhD",
     "Institute for Pure and Applied Knowledge (IPAK); independent",
     "@lifebiomedguru", "ipaknowledge.org; popularrationalism.substack.com",
     "N",
     "Autism epigenetics; vaccine epidemiology; environmental and genetic causes of autism",
     "Authored 'The Environmental and Genetic Causes of Autism' (Frye foreword). Bioinformatics PhD. Critical of CDC vaccine surveillance methodology.",
     "HYP-0044; HYP-0028; HYP-0072 (epigenetic canalization)",
     "Contested / mainstream-skeptical",
     6,
     "Independent researcher; left University of Pittsburgh 2009. Heavily advocacy-engaged. Primary research output mixed with frequent commentary.",
     "Medium",
     "Higher engagement-bait ratio than Frye/Adams. Filter for substantive paper-cites; ignore retweet-rage posts."),

    (9, "Brian Hooker", "PhD",
     "Children's Health Defense; Simpson University (formerly)",
     "@BrianHookerPhD", "childrenshealthdefense.org",
     "Y",
     "Vaccine-autism methodology critique; CDC Whistleblower (Thompson) documents",
     "Chemical engineering PhD. Parent of vaccine-injured son. Co-author atlas SRC-001434 (Hooker 2014 critique). Recipient of Thompson documents (atlas SRC-001419).",
     "HYP-0044; HYP-0068; SRC-001419 (Thompson docs); SRC-001434",
     "Contested / mainstream-skeptical",
     5,
     "His 2014 reanalysis paper (different from SRC-001434) was retracted; reduces methodological reliability score. CHD employee.",
     "Medium-Low",
     "Use for: Thompson document context, lab reasoning. DON'T use for: novel epidemiology (track record of methodological criticism)."),

    (10, "Toby Rogers", "PhD",
     "Brownstone Institute (Fellow); independent",
     "(less direct X presence; primary platform: Substack)", "tobyrogers.substack.com",
     "N",
     "Political economy of autism; regulatory capture; chronic illness epidemic in children",
     "PhD thesis 'The Political Economy of Autism' (U. Sydney). Senate testimony on autism. Substack-primary.",
     "Atlas concept-level: regulatory-capture framing for HYP-0044 contested status",
     "Legal / policy",
     7,
     "Political-economy lens, not biomedical. Substantive on regulatory critique; less primary biomedical research output.",
     "Medium-High",
     "Use for context on WHY mainstream evidence is mid (regulatory/funding asymmetry analysis), not for biomedical signal."),

    (11, "Stephanie Seneff", "PhD",
     "MIT CSAIL (senior research scientist; computer science background)",
     "(handle: limited public X; primary via lectures + papers)", "people.csail.mit.edu/seneff",
     "N",
     "Glyphosate-autism; sulfur metabolism; methionine synthase; vitamin D",
     "Multiple peer-reviewed papers proposing glyphosate-autism causal mechanism. Heavily criticized; mainstream genetic literacy projects rebut her central claims. Some primary glyphosate-biochemistry papers stand independently.",
     "HYP-0005 (glyphosate); MEC-0003 (methylation)",
     "Contested / mainstream-skeptical",
     5,
     "Computer scientist, not biomedical PhD. Made famous wrong prediction (50% autism by 2025). Heavy criticism for cherry-picking; some valid biochemical observations buried in larger contested claims.",
     "Low-Medium",
     "Filter heavily. Specific glyphosate-methionine biochemistry citations may be valid; broad apocalyptic predictions are unreliable."),

    (12, "Daniel Geschwind", "MD, PhD",
     "UCLA Center for Autism Research and Treatment",
     "(handle uncertain; minimal public X)", "geschwindlab.dgsom.ucla.edu",
     "N",
     "ASD genetics; SFARI Tier 1 gene network; PsychENCODE; transcriptomics",
     "Foundational ASD genetics researcher. Co-discovered numerous SFARI genes. Atlas connects via the 781 SFARI Tier 1+2 genes wired to HYP-0028.",
     "HYP-0028 (polygenic risk); MEC-0006 (synapse); the gene layer",
     "Mainstream academic",
     9,
     "Mainstream academic. NIH-funded. No advocacy alignment. Standard peer-review track.",
     "Very High",
     "Track via published papers, not Twitter. Mainstream-genetics anchor researcher."),

    (13, "Eric Hollander", "MD",
     "Albert Einstein College of Medicine; Montefiore",
     "(handle uncertain)", "einsteinmed.edu",
     "N",
     "Autism psychopharmacology; OCD-autism overlap; clinical trials infrastructure",
     "Heads autism trial group. Original studies on intranasal oxytocin and other behavioral interventions.",
     "Connects to MEC-0019 (vagus/autonomic); future oxytocin/vasopressin work in Session 3.5G",
     "Mainstream academic",
     7,
     "Pharma-trial-funded; standard mainstream academic conflict landscape.",
     "Medium-High",
     "Mainstream-academic signal; track via published clinical trials."),

    (14, "Aaron Siri", "JD",
     "Siri & Glimstad LLP; ICAN counsel",
     "@AaronSiriSG", "sirilabel.com",
     "N",
     "Vaccine-injury litigation; FOIA against FDA/CDC; vaccine schedule oversight",
     "Lead counsel on multiple FOIA suits that produced atlas-relevant primary documents (Verstraeten emails, Generation Zero, ACIP records). Author 'Vaccines, Amen!' critical of NCVIA 1986.",
     "Atlas SRC-001415 to SRC-001420 (FOIA documents) trace partly to ICAN-driven litigation",
     "Legal / policy",
     8,
     "Attorney; advocacy-aligned. NOT a biomedical researcher. Best signal: FOIA-released document announcements.",
     "High",
     "Follow specifically for FOIA-released CDC/FDA documents — those are tier-1 primary atlas evidence per CLAUDE.md §4."),

    (15, "Mary Holland", "JD",
     "Children's Health Defense (President); formerly NYU Law (17 yrs)",
     "(handle uncertain; primarily via CHD)", "childrenshealthdefense.org",
     "Y",
     "Vaccine policy law; legal scholarship on vaccine injury",
     "Parent of autistic child (developmental regression). NYU Law alum + 17-year faculty. Co-author of vaccine-injury legal-scholarship papers.",
     "Atlas concept-level: NCVIA 1986 liability shield context for HYP-0044",
     "Legal / policy",
     6,
     "CHD president; advocacy-aligned. Legal scholarship is substantive but not biomedical primary research.",
     "Medium-High",
     "Use for: vaccine-injury legal/policy framing. Don't use for: biomedical signal."),

    (16, "Robert F. Kennedy Jr.", "JD",
     "U.S. Department of Health and Human Services (Secretary, 2025-)",
     "(public X presence as Secretary)", "hhs.gov; childrenshealthdefense.org (former)",
     "N (extended family advocacy; not parent of an ASD child personally to my knowledge)",
     "Founded Children's Health Defense; current HHS Secretary; leads federal autism initiative",
     "Atlas connects via: HYP-0044 contested-positive cluster; SRC-001420 (ACIP Sep 2025); Hep B birth-dose policy review",
     "Legal / policy",
     5,
     "Now in policy role. Statements have political weight beyond their evidentiary basis. Often cites Generation Zero 11.35× as '10,000%' (rhetorical inflation per CLAUDE.md HYP-0066 description).",
     "Medium (for biomedical claims) / High (for policy news)",
     "Highest signal: actual HHS policy actions and announcements (these reshape research funding landscape). Lower signal: biomedical claims (frequently rhetorically inflated)."),

    (17, "Mary Beth Pfeiffer", "(journalist; no MD/PhD but sustained primary investigative work)",
     "Independent investigative journalist",
     "@marybethpf", "marybethpfeiffer.com",
     "N",
     "Lyme disease; autism; chronic illness; long COVID — investigative-journalism scale",
     "Author 'Lyme: The First Epidemic of Climate Change.' Atlantic, NYT, USA Today contributor. Long-form investigative work on autism prevalence.",
     "Atlas concept-level: contributes to HYP-0044 contested-evidence framing",
     "Journalist / parent-advocate",
     7,
     "Journalist, not researcher. Rigorous evidence-based reporting. Bias: critical of medical establishment.",
     "High (for journalism)",
     "Treat as investigative-journalism signal: she vets her sources well. Often surfaces FOIA documents and primary studies that mainstream coverage misses."),

    (18, "Polly Tommey", "(parent advocate; producer)",
     "CHD.TV (Director); VAXXED documentary producer",
     "(handle uncertain on X; primary platform CHD.TV)", "vaxxed.com; chd.tv",
     "Y",
     "Parent-perspective autism advocacy; vaccine-injury documentation",
     "Mother of Billy Tommey (vaccine injury → autism case). Founded The Autism File magazine. Produced VAXXED, VAXXED II, VAXXED III. Documentary-scale aggregation of parent-stories.",
     "Concept-level: contributes parent-experience evidence (anecdote tier per atlas; type=anecdote weight 0.15)",
     "Journalist / parent-advocate",
     4,
     "Strong advocacy alignment. Documentary content is anecdote-tier evidence per CLAUDE.md schema.",
     "Medium",
     "Aggregator of parent-experience anecdotes. Use for: subjective-response patterns. Don't use for: causal inference."),

    (19, "Peter A. McCullough", "MD, MPH",
     "McCullough Foundation; Truth For Health Foundation",
     "@P_McCulloughMD", "mcculloughfnd.org",
     "N",
     "Cardiology (primary); recently expanded to vaccine injury; planned autism case-series",
     "Cardiologist with extensive cardiology publication track. Recently launched McCullough Foundation autism case-series of regression-after-vaccination cases (planned, ongoing).",
     "HYP-0044 (concept-level support)",
     "Contested / mainstream-skeptical",
     6,
     "ABIM revoked board certifications 2025 for COVID misinformation. Cardiology track record substantial; autism work is recent and not yet methodologically validated.",
     "Medium",
     "Watch for: McCullough Foundation autism case-series outputs (could become primary evidence). Filter heavily for COVID-era engagement-bait posts."),

    (20, "William J. Walsh", "PhD",
     "Walsh Research Institute (founder)",
     "(no significant X presence; primary platform: institute + clinical staff)", "walshinstitute.org",
     "N",
     "Biotyping for autism (undermethylator/overmethylator/pyroluria/copper-zinc); Pfeiffer Treatment Center successor",
     "Walsh biotyping framework. Decades of clinical data (>30,000 patient records). Publishes through Walsh Research Institute. Less peer-reviewed RCT output (CLAUDE.md §10 notes the funding-asymmetry reason).",
     "Atlas Session 3.5D will formally ingest Walsh biotypes as phenotype subdivisions",
     "Functional medicine / clinical",
     6,
     "Clinical-experience-driven, lower peer-reviewed output. Per CLAUDE.md §10: don't dismiss for non-mainstream — funding asymmetry is real.",
     "Medium-High (for biotyping framework)",
     "Use for: methylation phenotype refinement. Track via Walsh Research Institute publications, not Twitter."),

    # Honorable mentions (21-25)
    (21, "Lucija Tomljenovic", "PhD",
     "Independent (formerly UBC)",
     "(handle uncertain)", "publication record",
     "N",
     "Aluminum vaccine adjuvants; Hill criteria ecological work",
     "Atlas SRC-001431 author. Tomljenovic-Shaw group. Heavily criticized by mainstream but published peer-reviewed research.",
     "HYP-0067; SRC-001431",
     "Contested / mainstream-skeptical",
     6,
     "Lost UBC affiliation due to advocacy concerns. Primary research findings stand independently.",
     "Medium",
     "Honorable mention; primary research is in atlas; less Twitter-active."),

    (22, "Russell Blaylock", "MD",
     "Theoretical Neuroscience Research; retired neurosurgeon",
     "(handle uncertain on X)", "russellblaylockmd.com",
     "N",
     "Immunoexcitotoxicity; microglial priming; vaccine-induced neuroinflammation",
     "Atlas SRC-001444 author. 'Immunoexcitotoxicity' framework concept (microglial-priming part now mainstream-validated by Bilbo, Tetreault, Vargas).",
     "MEC-0002 neuroinflammation; MEC-0005 microglial; SRC-001444",
     "Contested / mainstream-skeptical",
     5,
     "Alternative-medicine-aligned. Excitotoxin theorist. Mainstream-criticized but mechanism concept partially validated independently.",
     "Medium",
     "Honorable mention; framework conceptually valuable but advocacy-heavy."),

    (23, "Suzanne Goh", "MD",
     "Cortica Healthcare (Chief Medical Officer)",
     "(public via Cortica)", "corticacare.com",
     "N",
     "Clinical autism medicine; functional-medicine-inflected clinical practice",
     "Pediatric neurologist. Cortica integrates functional medicine + behavioral therapy + medical workup. Less peer-reviewed publication output but substantive clinical-protocol contribution.",
     "Connects to functional medicine intervention layer (Session 3.5B)",
     "Functional medicine / clinical",
     7,
     "Cortica is for-profit; financial alignment with services they provide.",
     "Medium-High",
     "Honorable mention; clinical protocol expert. Track via Cortica publications."),

    (24, "Pierre Kory", "MD",
     "FLCCC Alliance (President); Leading Edge Clinic",
     "@PierreKory", "flccc.net; legendsclinic.com",
     "N",
     "Critical care; vaccine injury; long COVID; PANS/PANDAS via FLCCC autism extension",
     "Pulmonary/critical care MD. FLCCC has expanded into PANS/PANDAS and vaccine-injury clinical care including autism-overlapping cases.",
     "Connects to PANS/PANDAS hypothesis (Session 3.5C)",
     "Contested / mainstream-skeptical",
     5,
     "Heavy COVID-era engagement; recently more autism-relevant. Filter heavily.",
     "Medium",
     "Honorable mention; PANS-treatment clinical experience could feed atlas Session 3.5C."),

    (25, "Martin Kulldorff", "PhD",
     "Hillsdale College Academy of Science (formerly Harvard biostatistics)",
     "@MartinKulldorff", "kulldorff.com",
     "N",
     "Vaccine surveillance biostatistics; epidemiological methodology; co-author Great Barrington Declaration",
     "Harvard biostatistics professor (forced out 2024). Designed CDC's vaccine adverse event detection algorithms. Highest credentialing in vaccine-surveillance methods.",
     "Concept-level: methodology critique relevant to HYP-0044 mainstream-null studies",
     "Mainstream academic",
     8,
     "Most credentialed biostatistician in this list. Career-damaged for COVID-era policy critique. Methodology critique stands on technical merits.",
     "Very High (for methodology)",
     "Honorable mention; technical methodology authority for vaccine-surveillance critique. Treat as gold-standard biostat reference."),
]

# === Build workbook ===
wb = Workbook()

# --- Sheet 1: Researchers ---
sh1 = wb.active
sh1.title = "Researchers"

headers1 = [
    "#", "Name", "Credentials", "Affiliation", "Twitter/X handle",
    "Other primary platform", "Parent of ASD child?",
    "Research focus", "Key contributions to autism science",
    "Atlas connections (HYP/MEC/INT/SRC)", "Tier",
    "Signal density (1-10)", "Bias considerations", "Reliability",
    "Notes / signal-vs-noise filter"
]

for col_idx, h in enumerate(headers1, 1):
    cell = sh1.cell(1, col_idx, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER

for row_idx, r in enumerate(researchers, 2):
    for col_idx, val in enumerate(r, 1):
        cell = sh1.cell(row_idx, col_idx, val)
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGN
        cell.border = BORDER
    # Tier color code (column K = 11)
    tier = r[10]
    if tier in TIER_FILLS:
        for col_idx in range(1, len(r)+1):
            sh1.cell(row_idx, col_idx).fill = TIER_FILLS[tier]

# Column widths
widths1 = [4, 22, 12, 35, 22, 28, 6, 35, 50, 35, 28, 8, 40, 12, 45]
for i, w in enumerate(widths1, 1):
    sh1.column_dimensions[get_column_letter(i)].width = w

# Freeze top row
sh1.freeze_panes = "A2"
sh1.row_dimensions[1].height = 36
for r in range(2, len(researchers)+2):
    sh1.row_dimensions[r].height = 100

# --- Sheet 2: Key Papers ---
sh2 = wb.create_sheet("Key Papers")
headers2 = ["Researcher", "Year", "PMID", "Journal", "Title (truncated)",
            "Atlas SRC if ingested", "Strength score", "Notes"]
for col_idx, h in enumerate(headers2, 1):
    cell = sh2.cell(1, col_idx, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER

# Key papers - one row per top paper per researcher (covers atlas-ingested + 1-2 honorable mentions)
key_papers = [
    ("Frye", 2018, "30097774", "Mol Psychiatry", "Folinic acid improves verbal communication in autism (RCT)", "SRC-000001", 0.5225, "Atlas calibration anchor RCT"),
    ("Frye", 2017, "28250023", "(varies)", "Cerebral folate receptor autoantibodies in autism", "SRC-000002", 0.5, "Foundational FOLR1 work"),
    ("Frye", 2021, "34834493", "Pediatr Rev", "Treatment of folate metabolism abnormalities in ASD", "SRC-000011", 0.5, "Atlas-ingested"),
    ("Rossignol", 2009, "(multiple)", "Med Hypotheses", "HBOT for autism — early protocol papers", "SRC related", 0.4, "Multiple HBOT papers"),
    ("Naviaux", 2017, "28503133", "Ann Clin Transl Neurol", "SAT1: low-dose suramin in autism (small open-label trial)", "(not yet ingested)", 0.4, "Suramin antipurinergic. Small N=10 but mechanistically novel"),
    ("Naviaux", 2014, "(multiple)", "Mitochondrion", "Cell danger response (CDR) framework", "concept-level in atlas", 0.6, "Foundational CDR theory"),
    ("Theoharides", 2016, "27272262", "Brain Behav Immun", "Brain mast cell activation in autism", "(not yet ingested)", 0.5, "Foundational MCAS-autism"),
    ("Hazan", 2024, "(varies)", "Microbial Cell Reports", "Bifidobacterium infantis and autism", "concept-level", 0.4, "Microbiome work"),
    ("Adams", 2017, "29076465", "Sci Rep", "Microbiota Transfer Therapy for autism (open-label, 2-year)", "(not yet ingested)", 0.5, "MTT foundational paper"),
    ("Adams", 2019, "30967259", "ASU News", "MTT 2-year follow-up (~50% symptom reduction sustained)", "(not yet ingested)", 0.5, "Long-term MTT outcomes"),
    ("Exley", 2018, "29161691", "J Trace Elem Med Biol", "Aluminum in brain tissue in autism", "(not yet ingested)", 0.4, "Mochizuki-Exley 2018 — high Al in ASD brain"),
    ("Lyons-Weiler", 2020, "33218038", "Int J Environ Res Public Health", "Relative incidence of office visits and cumulative diagnoses along axis of vaccination", "(not yet ingested)", 0.25, "Correction issued; controversial methodology"),
    ("Hooker", 2014, "24995277", "Biomed Res Int", "Methodological issues and evidence of malfeasance...", "SRC-001434", 0.15, "Atlas-ingested at low strength per CLAUDE.md §3"),
    ("Rogers", 2019, "(thesis)", "U. Sydney", "The Political Economy of Autism (PhD thesis)", "concept-level", 0.3, "Political economy framework"),
    ("Seneff", 2013, "(varies)", "Entropy", "Glyphosate and modern diseases", "(not ingested)", 0.15, "Heavily criticized; hypothesis-tier"),
    ("Geschwind", 2018, "30382198", "Science", "Transcriptome-wide isoform-level dysregulation in ASD", "(connects via gene layer)", 0.6, "PsychENCODE-tier mainstream genomics"),
    ("Hollander", 2014, "23548579", "Mol Autism", "Intranasal oxytocin in autism", "(connects to oxytocin layer)", 0.4, "Pioneer clinical trial"),
    ("Walsh", 2020, "(book)", "Nutrient Power Autism", "Nutrient Power for autism (clinical biotyping)", "Session 3.5D plan", 0.3, "Clinical-experience aggregation"),
    ("Tomljenovic", 2011, "22099159", "J Inorg Biochem", "Aluminum vaccine adjuvants and autism (Hill criteria)", "SRC-001431", 0.30, "Atlas-ingested"),
    ("Blaylock", 2008, "19043938", "Altern Ther Health Med", "Immunoexcitotoxicity central mechanism in ASD", "SRC-001444", 0.10, "Atlas-ingested at low strength"),
    ("Goh", 2024, "(Cortica)", "Cortica protocols", "Whole-child autism medicine (clinical-protocol)", "concept-level", 0.35, "Clinical protocol authority"),
    ("McCullough", 2025, "(forthcoming)", "McCullough Foundation", "Autism regression-after-vaccination case series (planned)", "(not yet)", "TBD", "Methodology not yet established"),
    ("Kulldorff", 2003, "(multiple)", "Pharmacoepidemiol Drug Saf", "VSD methodology / scan statistic for vaccine surveillance", "concept-level", 0.6, "Methodology authority for vaccine epidemiology"),
]

for row_idx, paper in enumerate(key_papers, 2):
    for col_idx, val in enumerate(paper, 1):
        cell = sh2.cell(row_idx, col_idx, val)
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGN
        cell.border = BORDER

widths2 = [22, 8, 14, 22, 50, 22, 14, 35]
for i, w in enumerate(widths2, 1):
    sh2.column_dimensions[get_column_letter(i)].width = w
sh2.freeze_panes = "A2"

# --- Sheet 3: Signal vs Noise ---
sh3 = wb.create_sheet("Signal vs Noise Guide")
headers3 = ["Researcher", "Signal density (1-10)",
            "What to weight as SIGNAL",
            "What to filter as NOISE", "Practical filter rule"]
for col_idx, h in enumerate(headers3, 1):
    cell = sh3.cell(1, col_idx, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER

signal_noise = [
    ("Frye", 9,
     "Posts citing his own RCTs; FOLR1 + leucovorin discussion; cerebral folate deficiency phenotype refinement; PANS treatment protocols",
     "Almost no noise — his Twitter is research-focused",
     "Default: read everything"),
    ("Rossignol", 7,
     "Posts on mito biomarker stratification; HBOT response prediction",
     "Limited Twitter; minimal noise",
     "Read all (low volume)"),
    ("Naviaux", 8,
     "Suramin trial updates; CDR theory papers; metabolomics findings",
     "Rare commentary; signal-pure",
     "Read all (low volume)"),
    ("Theoharides", 8,
     "Mast cell biology in autism; flavonoid mechanism",
     "Some commercial luteolin promotion (NeuroProtek)",
     "Filter for non-commercial mast cell science"),
    ("Hazan", 7,
     "Bifidobacterium infantis depletion findings; FMT protocols; specific microbiome findings",
     "Some commercial-adjacent posts (her clinic services)",
     "Filter: keep posts citing primary papers; skip clinic-promotion"),
    ("Adams", 9,
     "Almost everything — high signal density, low Twitter activity",
     "Limited X presence",
     "Subscribe via ASU page rather than Twitter"),
    ("Exley", 8,
     "Aluminum biochemistry findings; Mochizuki-Exley brain Al data",
     "Career-loss content; advocacy fundraising",
     "Filter: keep biochemistry signal; skip personal-circumstance content"),
    ("Lyons-Weiler", 6,
     "His own peer-reviewed paper announcements; specific epigenetic/methylation findings",
     "Frequent retweets of advocacy content; engagement-bait responses",
     "Filter heavily: keep posts that link to a PMID; skip pure commentary"),
    ("Hooker", 5,
     "Thompson document references; specific FOIA-document announcements",
     "His 2014 retraction reduces methodology trust; CHD-aligned advocacy",
     "Filter: trust him on FOIA documents and Thompson context; don't trust his epidemiology reanalyses"),
    ("Rogers", 7,
     "Substack long-form regulatory-economy analysis; Senate testimony references",
     "Less noise but limited biomedical signal",
     "Read substack long-form; ignore Twitter brevity"),
    ("Seneff", 5,
     "Specific glyphosate-methionine biochemistry citations; sulfur metabolism observations",
     "Apocalyptic predictions; over-broad causation claims",
     "Filter heavily: keep narrow biochemistry; skip prediction posts"),
    ("Geschwind", 9,
     "Lab paper announcements; PsychENCODE updates; SFARI-tier gene findings",
     "Almost no Twitter content",
     "Track via Pubmed alerts on his lab"),
    ("Hollander", 7,
     "Trial recruitment; clinical trial outcomes; psychopharmacology updates",
     "Limited Twitter activity",
     "Track via clinicaltrials.gov + lab page"),
    ("Siri", 8,
     "FOIA-released document announcements; legal filings producing primary documents",
     "Some advocacy framing of legal wins",
     "Highest priority for FOIA document signals; filter for actual document content"),
    ("Holland", 6,
     "Legal/policy framing of vaccine injury; CHD documentation",
     "CHD-aligned advocacy; not biomedical",
     "Use for legal/policy context; not biomedical claims"),
    ("Kennedy", 5,
     "HHS policy actions; ACIP review announcements; federal-level autism initiative",
     "Frequent rhetorical inflation of biomedical claims (e.g., '10,000% increase')",
     "Trust for: federal policy news. Don't trust for: biomedical specifics."),
    ("Pfeiffer", 7,
     "Investigative-journalism long-form pieces; FOIA document surfacing",
     "Limited noise (rigorous reporter)",
     "High trust for sourced reporting"),
    ("Tommey", 4,
     "Parent-experience aggregation; documentary content",
     "Heavy advocacy framing; anecdote-tier evidence",
     "Use sparingly; subjective-response context only"),
    ("McCullough", 6,
     "McCullough Foundation autism case series (when published); cardiology fundamentals",
     "COVID-era positions still dominate his feed",
     "Watch specifically for autism Foundation outputs; filter COVID content"),
    ("Walsh", 6,
     "Walsh biotyping framework; clinical-experience aggregation",
     "Limited Twitter; long-form via institute publications",
     "Use Walsh Research Institute publications; not Twitter"),
    ("Tomljenovic", 6,
     "Aluminum adjuvant primary research findings",
     "Career-loss content; less recent output",
     "Track historical primary research; less Twitter activity"),
    ("Blaylock", 5,
     "Immunoexcitotoxicity framework references",
     "Heavy advocacy/alt-med engagement",
     "Use framework references; filter advocacy"),
    ("Goh", 7,
     "Cortica clinical protocols; whole-child autism medicine",
     "Cortica-commercial alignment",
     "Track via Cortica protocols; filter commercial framing"),
    ("Kory", 5,
     "FLCCC PANS/autism extension protocols",
     "Heavy COVID-era residual content",
     "Filter heavily for autism-specific posts"),
    ("Kulldorff", 8,
     "Vaccine-surveillance methodology critique; biostatistics rigor",
     "Some COVID-era policy positioning",
     "Highest trust for biostatistical methodology critique"),
]

for row_idx, sn in enumerate(signal_noise, 2):
    for col_idx, val in enumerate(sn, 1):
        cell = sh3.cell(row_idx, col_idx, val)
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGN
        cell.border = BORDER

widths3 = [22, 8, 50, 45, 40]
for i, w in enumerate(widths3, 1):
    sh3.column_dimensions[get_column_letter(i)].width = w
sh3.freeze_panes = "A2"
sh3.row_dimensions[1].height = 36
for r in range(2, len(signal_noise)+2):
    sh3.row_dimensions[r].height = 60

# --- Sheet 4: Methodology + atlas linkage notes ---
sh4 = wb.create_sheet("Methodology + Atlas Linkage")
sh4.column_dimensions['A'].width = 110

methodology_notes = [
    ("Autism Researchers Tracker — Methodology Notes",),
    ("",),
    (f"Generated: {datetime.now().strftime('%Y-%m-%d')}",),
    ("",),
    ("PURPOSE",),
    ("Track 25 autism researchers + parent-advocates whose primary research output the atlas should ingest.",),
    ("Per CLAUDE.md mainstream-skeptical principle, the list spans mainstream academic, functional medicine,",),
    ("contested-evidence, and parent-researcher tiers. Their PEER-REVIEWED OUTPUT is the signal; Twitter is",),
    ("downstream commentary that points toward what they're working on.",),
    ("",),
    ("SIGNAL DENSITY SCALE (1-10)",),
    ("9-10: Almost all public output is research-grounded (e.g., Frye, Adams, Geschwind)",),
    ("7-8: Mostly research-grounded with occasional commentary (e.g., Naviaux, Theoharides, Kulldorff)",),
    ("5-6: Mixed; signal embedded in commentary (e.g., Lyons-Weiler, Walsh, Holland)",),
    ("3-4: Heavy advocacy/engagement framing (e.g., Tommey)",),
    ("1-2: Pure advocacy/political content (none on this list at this tier)",),
    ("",),
    ("RELIABILITY TIERS (per CLAUDE.md source-quality framework)",),
    ("Very High: Mainstream academic with clean track record (Geschwind, Adams, Kulldorff)",),
    ("High: Functional medicine with strong primary research (Frye, Naviaux, Theoharides)",),
    ("Medium-High: Some mixed elements but substantive primary work (Hazan, Exley, Pfeiffer)",),
    ("Medium: Mixed; needs filtering (Lyons-Weiler, Rogers, Walsh, Holland, Seneff, Goh)",),
    ("Medium-Low: Strong contested-evidence advocacy alignment (Hooker, McCullough, Blaylock, Kory)",),
    ("",),
    ("HOW TO USE THIS TRACKER",),
    ("1. For each researcher, set up a PubMed alert on their name + 'autism' or 'ASD'",),
    ("2. Set up Substack subscriptions for those whose primary platform is Substack",),
    ("3. When they tweet about a new paper, the paper itself goes into the atlas via run_ingest.py pmid",),
    ("4. Per CLAUDE.md, the atlas weights based on study design + source type, NOT author popularity",),
    ("5. Researchers with high signal density + high reliability are highest priority for ingestion",),
    ("",),
    ("PARENT-OF-AUTISTIC-CHILD STATUS",),
    ("3 confirmed: Adams (autistic adult daughter), Hooker (vaccine-injured son), Tommey (autistic son)",),
    ("Holland: parent of autistic child (developmental regression catalyst)",),
    ("RFK Jr.: extended family advocacy; not parent of an autistic child personally to my knowledge",),
    ("",),
    ("ATLAS CALIBRATION",),
    ("INT-0001 Leucovorin = 83.20 CSRS (calibration anchor)",),
    ("Calibration anchor researcher: Frye (Frye/Slattery FOLR1 work)",),
    ("",),
    ("LIMITATIONS",),
    ("1. Twitter handles for some researchers may have changed or be private",),
    ("2. Twitter content itself requires authentication to read; this tracker focuses on PUBLISHED work",),
    ("3. Some researchers have moved to Substack/BlueSky/Mastodon since X transition",),
    ("4. Verify handles before subscribing (search for active accounts)",),
    ("",),
    ("NEXT STEPS PER ATLAS ROADMAP (CLAUDE.md Session 3.5+)",),
    ("- Session 3.5A: ingest functional medicine biomarkers; many of these researchers' work feeds in",),
    ("- Session 3.5B: ingest functional medicine interventions",),
    ("- Session 3.5C: PANS/MCAS hypothesis families (Theoharides primary)",),
    ("- Session 3.5D: Walsh/Pfeiffer phenotype refinement (Walsh primary)",),
    ("- Session 3.5E: named-protocol entities (Frye, Walsh, Bock, Klinghardt, MAPS)",),
]

for row_idx, line in enumerate(methodology_notes, 1):
    cell = sh4.cell(row_idx, 1, line[0])
    if row_idx == 1:
        cell.font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    elif line[0] in ("PURPOSE", "SIGNAL DENSITY SCALE (1-10)", "RELIABILITY TIERS (per CLAUDE.md source-quality framework)",
                      "HOW TO USE THIS TRACKER", "PARENT-OF-AUTISTIC-CHILD STATUS", "ATLAS CALIBRATION",
                      "LIMITATIONS", "NEXT STEPS PER ATLAS ROADMAP (CLAUDE.md Session 3.5+)"):
        cell.font = Font(name='Arial', size=11, bold=True, color='1F4E79')
    else:
        cell.font = Font(name='Arial', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Save
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheet 1 'Researchers': {len(researchers)} rows")
print(f"Sheet 2 'Key Papers': {len(key_papers)} rows")
print(f"Sheet 3 'Signal vs Noise Guide': {len(signal_noise)} rows")
print(f"Sheet 4 'Methodology + Atlas Linkage': overview")
