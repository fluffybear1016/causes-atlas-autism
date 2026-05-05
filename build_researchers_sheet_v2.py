#!/usr/bin/env python3
"""Autism_Researchers.xlsx v2 — substantially enriched.

New in v2:
  - Sheet 1 'Researchers': baseline tracker
  - Sheet 2 'Recent Publications (2023-2026)': PubMed-pulled recent
    autism-relevant papers per researcher
  - Sheet 3 'Recent Substack/Public Findings': non-PubMed recent activity
    (Substack posts, reports, public commentary)
  - Sheet 4 'Key Atlas-Ingested Papers': papers already in atlas with SRC IDs
  - Sheet 5 'Signal vs Noise Guide': quality assessment per researcher
  - Sheet 6 'Methodology + Atlas Linkage': overview

Recent findings ingested via PubMed esearch + esummary, plus targeted
WebSearch for those without PubMed activity (Substack-primary researchers).
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

OUT = "Autism_Researchers.xlsx"

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', start_color='1F4E79')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
BODY_FONT = Font(name='Arial', size=10)
BODY_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
THIN = Side(border_style='thin', color='888888')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TIER_FILLS = {
    'Functional medicine / clinical': PatternFill('solid', start_color='E2EFDA'),
    'Contested / mainstream-skeptical': PatternFill('solid', start_color='FFF2CC'),
    'Mainstream academic': PatternFill('solid', start_color='D9E1F2'),
    'Legal / policy': PatternFill('solid', start_color='FCE4D6'),
    'Journalist / parent-advocate': PatternFill('solid', start_color='F4CCCC'),
}

# Load PubMed pulls done earlier
try:
    recent_pubs = json.load(open("v2.0.1_proposed/researcher_recent_pubs.json"))
except Exception:
    recent_pubs = {}


# === Researchers (main) ===
researchers = [
    (1, "Richard E. Frye", "MD, PhD",
     "Rossignol Medical Center; Autism Discovery & Treatment Foundation",
     "@AutismEnergetic", "drfryemdphd.com",
     "N",
     "Cerebral folate deficiency; mitochondrial autism; FOLR1 autoantibodies; methylation; PANS",
     "300+ peer-reviewed publications. Lead investigator on multiple leucovorin RCTs. Frye/Slattery FOLR1 protocol. Atlas calibration anchor scientist (INT-0001 leucovorin = 83.20 CSRS).",
     "INT-0001; HYP-0001; PHE-0001; SRC-000001; SRC-000011-26 (multiple)",
     "Functional medicine / clinical",
     9, "Functional medicine practitioner; criticized by mainstream as advocacy-leaning, but his peer-reviewed track record is solid. Owns clinic that uses interventions he researches.",
     "High",
     "STRONGEST signal. Active 2025-2026 publishing on gut-brain inflammation, transgenerational FOLR1 autoantibodies, mito transcriptomic signatures."),

    (2, "Daniel A. Rossignol", "MD",
     "Rossignol Medical Center", "(handle uncertain)", "rossignolmedicalcenter.com",
     "N",
     "HBOT for autism; mito autism; functional medicine; methyl-B12 protocols",
     "55+ peer-reviewed papers. Long-term Frye collaborator. Founded Rossignol Medical Center. Recent: 2025 paper on early ASD biomarkers; 2024 mitochondrial biomarkers paper.",
     "INT-0092 (HBOT); HYP-0001; multiple SRC",
     "Functional medicine / clinical",
     7,
     "Same financial overlap concern as Frye. Less Twitter-active.",
     "High",
     "Active 2024-2025 publishing on mito biomarkers, early biomarkers, leucovorin response."),

    (3, "Robert K. Naviaux", "MD, PhD",
     "UCSD Mitochondrial and Metabolic Disease Center",
     "(rare X use)", "naviauxlab.ucsd.edu",
     "N",
     "Cell danger response (CDR); suramin antipurinergic; metabolomics; 3-hit metabolic signaling model",
     "Originated CDR framework. SAT1 trial (suramin in autism, 2017). 2024 paper: metabolic network analysis of pre-ASD newborns. 2026: '3-hit metabolic signaling model' paper in Mitochondrion + Autism Research.",
     "MEC-0010 (mito); HYP-0001 (CDR); future suramin entry",
     "Mainstream academic",
     8,
     "Academic, not advocacy-aligned.",
     "High",
     "Latest framework: '3-hit metabolic signaling model' (PMID 41902612, 2026, Autism Research). Major framework update — pre-ASD biomarker work."),

    (4, "Theoharis C. Theoharides", "MD, PhD",
     "Tufts (emeritus); Institute for Neuro-Immune Medicine, Nova Southeastern",
     "(rare X)", "mastcellmaster.com",
     "N",
     "Mast cell activation in autism; brain-mast-cell axis; flavonoid therapeutics (luteolin)",
     "h-index 113. 2024: hippocampal molecular profiling in autism (Mol Psychiatry). 2024: ochratoxin (mycotoxin) → mast cell IL-1β/IL-18/CXCL8 release.",
     "MEC-0017; planned MCAS HYP (Session 3.5C); INT-0029 quercetin",
     "Functional medicine / clinical",
     8,
     "Patents on luteolin (NeuroProtek); financial interest disclosed.",
     "High",
     "Mainstream-published primary work. 2024-2025 papers feeding directly into atlas's planned MCAS hypothesis."),

    (5, "Sabine Hazan", "MD",
     "ProgenaBiome; Microbiome Research Foundation",
     "@SabinehazanMD", "progenabiome.com",
     "N",
     "Gut microbiome in autism; Bifidobacterium infantis; FMT; MTT",
     "2024: J Med Cases paper on gut microbiome composition + clinical symptoms post-MTT (correction issued same year).",
     "MEC-0008; HYP-0007; HYP-0056; INT-0025; INT-0076; INT-0077",
     "Functional medicine / clinical",
     7,
     "ProgenaBiome owner; financial interest in microbiome interventions.",
     "Medium-High",
     "Filter for posts citing primary papers vs commercial promotion."),

    (6, "James B. Adams", "PhD",
     "Arizona State University; Autism Nutrition Research Center",
     "(less X-active)", "autismnrc.org",
     "Y",
     "MTT; ASU autism nutrition; B-vitamins; carnitine; PANS",
     "150+ peer-reviewed papers. Adult autistic daughter is research catalyst. 2025 papers: PANS-Recurrent Outcome Survey validation (PRO-S); para-cresol elevation review; 2024: indoxyl sulfate review; bimodal Candida distribution in ASD; 2023: 13 therapeutic diets effectiveness ratings.",
     "INT-0076; HYP-0007; HYP-0056; broad nutritional interventions",
     "Functional medicine / clinical",
     9,
     "Academic with parental motivation. No major financial conflicts.",
     "Very High",
     "HIGHEST reliability tier. 2024-2025 substantive output on PANS validation, p-cresol, Candida, indoxyl sulfate. Direct atlas Session 3.5A biomarker layer feed."),

    (7, "Christopher Exley", "PhD",
     "Independent (formerly Keele Univ; forced out 2021)",
     "(Substack primary)", "drchristopherexley.substack.com",
     "N",
     "Aluminum biochemistry; aluminum in autism brain; aluminum adjuvants",
     "Mochizuki-Exley 2018 (high Al in autism brain) is the foundational paper. 2024-2025 Substack series: 'Aluminium and Autism: The Evidence' (July 2025), 'Unequivocal', 'Aluminium, Autism and Epilepsy', 'Tylenol, Autism and All That'.",
     "HYP-0067; MEC-0032; SRC-001431",
     "Contested / mainstream-skeptical",
     8,
     "Career destroyed for maintaining position. Primary findings replicated. Substack-primary now.",
     "Medium-High",
     "2025 Substack output sustained. 'Causal role for aluminium in autism' is his stated position."),

    (8, "James Lyons-Weiler", "PhD",
     "IPAK; independent",
     "@lifebiomedguru", "popularrationalism.substack.com",
     "N",
     "Autism epigenetics; vaccine epidemiology; environmental + genetic causes",
     "Authored 'Environmental and Genetic Causes of Autism' (Frye foreword). Active Substack 2025: 'CDC's Autism Reversal' (Nov 2025) covering CDC website language change, Atlantic rebuttal, Autism Health Summit April 2025.",
     "HYP-0044; HYP-0028; HYP-0072 (epigenetic canalization)",
     "Contested / mainstream-skeptical",
     6,
     "Independent; left Pittsburgh 2009. Heavy advocacy engagement.",
     "Medium",
     "Filter heavily. Most recent: CDC removed 'Vaccines Do Not Cause Autism' language Nov 2025 (per his reporting)."),

    (9, "Brian Hooker", "PhD",
     "Children's Health Defense (Chief Scientific Officer)",
     "@BrianHookerPhD", "childrenshealthdefense.org",
     "Y",
     "Vaccine-autism methodology critique; CDC Whistleblower (Thompson) documents",
     "Chemical engineering PhD. Parent of vaccine-injured son. Atlas SRC-001434 (2014 critique). Recipient of Thompson docs (atlas SRC-001419). No recent autism PubMed.",
     "HYP-0044; HYP-0068; SRC-001419; SRC-001434",
     "Contested / mainstream-skeptical",
     5,
     "2014 reanalysis paper retracted. CHD employee.",
     "Medium-Low",
     "Trust on FOIA documents. Methodologically less reliable for novel epidemiology."),

    (10, "Toby Rogers", "PhD",
     "Brownstone Institute (Fellow); independent",
     "(less X)", "tobyrogers.substack.com",
     "N",
     "Political economy of autism; regulatory capture; chronic illness epidemic",
     "PhD thesis 'Political Economy of Autism' (U. Sydney). Senate testimony. May 2025 Substack: 'Mapping the entire field of autism causation studies.' Sept 2025 Senate hearing testimony.",
     "Concept-level: regulatory-capture framing for HYP-0044",
     "Legal / policy",
     7,
     "Political-economy lens, not biomedical.",
     "Medium-High",
     "May 2025 'Mapping the entire field of autism causation' is a pseudo-meta-analysis worth ingesting concept-by-concept."),

    (11, "Stephanie Seneff", "PhD",
     "MIT CSAIL (computer science background)",
     "(limited public X)", "people.csail.mit.edu/seneff",
     "N",
     "Glyphosate-autism; sulfur metabolism; methylation; vitamin D",
     "2024: 'Is autism a PIN1 deficiency syndrome?' (J Neurochem). Heavily criticized; some primary biochemistry observations valid.",
     "HYP-0005 (glyphosate); MEC-0003",
     "Contested / mainstream-skeptical",
     5,
     "Computer scientist, not biomedical PhD. Famous wrong prediction (50% autism by 2025).",
     "Low-Medium",
     "PIN1 hypothesis 2024 paper is recent novel framework. Filter heavily."),

    (12, "Daniel H. Geschwind", "MD, PhD",
     "UCLA Center for Autism Research and Treatment",
     "(minimal X)", "geschwindlab.dgsom.ucla.edu",
     "N",
     "ASD genetics; SFARI gene network; PsychENCODE; transcriptomics",
     "2026: Nature paper on developmental convergence/divergence in human stem cell models of autism. 2025: Nature paper polygenic and developmental profiles by age at diagnosis. 2025 medRxiv: African American autism genetics; CNVs; psychiatric disorder convergence.",
     "HYP-0028 (polygenic); MEC-0006 (synapse); the gene layer",
     "Mainstream academic",
     9,
     "Mainstream academic. NIH-funded.",
     "Very High",
     "Multiple Nature papers 2025-2026. ABSORB into atlas — these are tier-1 mainstream genomics."),

    (13, "Eric Hollander", "MD",
     "Albert Einstein College of Medicine; Montefiore",
     "(handle uncertain)", "einsteinmed.edu",
     "N",
     "Autism psychopharmacology; OCD-autism; oxytocin/vasopressin; clinical trials",
     "2024: oxytocin in neurodevelopmental disorders (Pharmacology & Therapeutics); cognitive inflexibility + immunome biomarkers in autism. 2023: vasopressin V1a receptor antagonist clinical trials predictors.",
     "Connects to oxytocin/vasopressin layer (Session 3.5G)",
     "Mainstream academic",
     7,
     "Pharma-trial-funded.",
     "Medium-High",
     "Track recent oxytocin/vasopressin clinical trial outcomes — directly feeds Session 3.5G peptide layer."),

    (14, "Aaron Siri", "JD",
     "Siri & Glimstad LLP; ICAN counsel",
     "@AaronSiriSG", "sirilabel.com",
     "N",
     "Vaccine-injury litigation; FOIA against FDA/CDC; vaccine schedule oversight",
     "Lead counsel on FOIA suits producing atlas-relevant documents (Verstraeten emails, Generation Zero, ACIP records). 'Vaccines, Amen!' book on NCVIA 1986.",
     "Atlas SRC-001415 to SRC-001420 trace partly to ICAN-driven litigation",
     "Legal / policy",
     8,
     "Attorney; advocacy-aligned. NOT biomedical researcher.",
     "High",
     "Highest priority for FOIA-released document signals."),

    (15, "Mary Holland", "JD",
     "CHD President; formerly NYU Law (17 yrs)",
     "(handle uncertain)", "childrenshealthdefense.org",
     "Y",
     "Vaccine policy law; legal scholarship on vaccine injury",
     "Parent of autistic child. NYU Law alum. Co-author vaccine-injury legal-scholarship papers.",
     "Concept-level: NCVIA 1986 liability shield context",
     "Legal / policy",
     6,
     "CHD president; advocacy-aligned.",
     "Medium-High",
     "Use for legal/policy framing; not biomedical."),

    (16, "Robert F. Kennedy Jr.", "JD",
     "U.S. HHS Secretary (2025-)",
     "(public X as Secretary)", "hhs.gov",
     "N (extended family advocacy)",
     "Founded CHD; current HHS Secretary; leads federal autism initiative",
     "Atlas via HYP-0044 cluster + SRC-001420 ACIP review.",
     "Multiple atlas connections via HHS policy actions",
     "Legal / policy",
     5,
     "Statements have political weight beyond evidentiary basis. Frequently cites Generation Zero 11.35× as '10,000%'.",
     "Medium / High (for policy)",
     "Policy actions reshape research funding landscape. Track for ACIP votes, MAHA initiative, federal autism announcements."),

    (17, "Mary Beth Pfeiffer", "(journalist; sustained primary investigative work)",
     "Independent investigative journalist",
     "@marybethpf", "marybethpfeiffer.com",
     "N",
     "Lyme disease; autism; chronic illness; long COVID — investigative journalism",
     "Author 'Lyme: The First Epidemic of Climate Change.' Atlantic, NYT, USA Today. Long-form investigative work on autism prevalence.",
     "Concept-level: contributes to HYP-0044 framing",
     "Journalist / parent-advocate",
     7,
     "Journalist, not researcher. Rigorous evidence-based reporting.",
     "High (for journalism)",
     "Surfaces FOIA documents and primary studies mainstream coverage misses."),

    (18, "Polly Tommey", "(parent advocate; producer)",
     "CHD.TV (Director); VAXXED documentary producer",
     "(handle uncertain)", "vaxxed.com; chd.tv",
     "Y",
     "Parent-perspective autism advocacy; vaccine-injury documentation",
     "Mother of Billy Tommey (vaccine injury → autism case). Founded The Autism File magazine. Produced VAXXED, VAXXED II, VAXXED III.",
     "Concept-level: parent-experience evidence (anecdote tier)",
     "Journalist / parent-advocate",
     4,
     "Strong advocacy alignment. Documentary content is anecdote-tier per atlas schema.",
     "Medium",
     "Use for: subjective-response patterns. Don't use for: causal inference."),

    (19, "Peter A. McCullough", "MD, MPH",
     "McCullough Foundation; Truth For Health Foundation",
     "@P_McCulloughMD", "mcculloughfnd.org",
     "N",
     "Cardiology (primary); recently expanded to vaccine injury; autism case-series",
     "Cardiologist with extensive cardiology track. October 2025: McCullough Foundation Report on Determinants of ASD published on Zenodo (DOI 10.5281/zenodo.17451259) — review of 136 vaccine-autism studies; 107 inferred possible link, 29 found neutral. Planned large-scale case-series of regression-after-vaccination cases.",
     "HYP-0044 concept-level support; future SRC for Zenodo report",
     "Contested / mainstream-skeptical",
     6,
     "ABIM revoked board certifications 2025 for COVID misinformation. Zenodo report not peer-reviewed (preprint-tier).",
     "Medium",
     "Watch for: McCullough Foundation case-series outputs. Filter heavily for COVID-era engagement-bait."),

    (20, "William J. Walsh", "PhD",
     "Walsh Research Institute (founder)",
     "(no significant X)", "walshinstitute.org",
     "N",
     "Biotyping for autism (undermethylator/overmethylator/pyroluria/copper-zinc); Pfeiffer Treatment Center successor",
     "Walsh biotyping framework. Decades of clinical data (>30,000 patient records). 'Nutrient Power' book. Less peer-reviewed RCT output (CLAUDE.md §10 — funding asymmetry reason).",
     "Atlas Session 3.5D will ingest Walsh biotypes",
     "Functional medicine / clinical",
     6,
     "Clinical-experience-driven, lower peer-reviewed output.",
     "Medium-High (for biotyping)",
     "Use for methylation phenotype refinement."),

    (21, "Lucija Tomljenovic", "PhD",
     "Independent (formerly UBC Neural Dynamics Research)",
     "(handle uncertain)", "publication record",
     "N",
     "Aluminum vaccine adjuvants; Hill criteria ecological work",
     "Atlas SRC-001431 author. Tomljenovic-Shaw group. Heavily criticized but published peer-reviewed research.",
     "HYP-0067; SRC-001431",
     "Contested / mainstream-skeptical",
     6,
     "Lost UBC affiliation due to advocacy concerns.",
     "Medium",
     "Honorable mention; primary research in atlas; less recent."),

    (22, "Russell L. Blaylock", "MD",
     "Theoretical Neuroscience Research; retired neurosurgeon",
     "(handle uncertain)", "russellblaylockmd.com",
     "N",
     "Immunoexcitotoxicity; microglial priming; vaccine-induced neuroinflammation",
     "Atlas SRC-001444 author. 'Immunoexcitotoxicity' framework concept (microglial-priming part now mainstream-validated).",
     "MEC-0002; MEC-0005; SRC-001444",
     "Contested / mainstream-skeptical",
     5,
     "Alternative-medicine-aligned. Excitotoxin theorist.",
     "Medium",
     "Honorable mention; framework conceptually valuable but advocacy-heavy."),

    (23, "Suzanne Goh", "MD",
     "Cortica Healthcare (Chief Medical Officer)",
     "(public via Cortica)", "corticacare.com",
     "N",
     "Clinical autism medicine; functional-medicine clinical practice",
     "Pediatric neurologist. Cortica integrates functional medicine + behavioral therapy. 2025 paper on penetrance estimates of CNVs (Genetics in Medicine).",
     "Connects to functional medicine intervention layer (Session 3.5B)",
     "Functional medicine / clinical",
     7,
     "Cortica is for-profit; financial alignment.",
     "Medium-High",
     "Honorable mention; clinical protocol expert."),

    (24, "Pierre Kory", "MD",
     "FLCCC Alliance (President); Leading Edge Clinic",
     "@PierreKory", "flccc.net; legendsclinic.com",
     "N",
     "Critical care; vaccine injury; long COVID; PANS/PANDAS via FLCCC autism extension",
     "FLCCC has expanded into PANS/PANDAS and vaccine-injury clinical care.",
     "Connects to PANS/PANDAS hypothesis (Session 3.5C)",
     "Contested / mainstream-skeptical",
     5,
     "Heavy COVID-era engagement; recently more autism-relevant.",
     "Medium",
     "Honorable mention; PANS-treatment clinical experience."),

    (25, "Martin Kulldorff", "PhD",
     "Hillsdale College Academy of Science (formerly Harvard biostatistics)",
     "@MartinKulldorff", "kulldorff.com",
     "N",
     "Vaccine surveillance biostatistics; epidemiological methodology; co-author Great Barrington Declaration",
     "Designed CDC's vaccine adverse event detection algorithms. Highest credentialing in vaccine-surveillance methods.",
     "Concept-level: methodology critique relevant to HYP-0044",
     "Mainstream academic",
     8,
     "Career-damaged for COVID-era policy critique. Methodology critique stands on technical merits.",
     "Very High (for methodology)",
     "Honorable mention; gold-standard biostat reference for vaccine-surveillance critique."),
]

# === Recent Publications (PubMed-pulled) ===
# Map researcher short name -> list of recent papers
recent_pubs_table = []
researcher_short_names = ["Frye","Rossignol","Naviaux","Theoharides","Hazan","Adams","Exley",
                           "Lyons-Weiler","Hooker","Rogers","Seneff","Geschwind","Hollander",
                           "Tomljenovic","Blaylock","Goh","Kory","Kulldorff","McCullough","Walsh"]
for short in researcher_short_names:
    pubs = recent_pubs.get(short, [])
    for p in pubs[:6]:
        recent_pubs_table.append((short, p["year"], p["pmid"], p["journal"][:50],
                                   p["title"][:140], p.get("first_author","")))

# === Recent Substack/Public Findings (non-PubMed) ===
substack_findings = [
    ("Frye", "drfryemdphd.com", "2025",
     "Cerebral folate deficiency clinical guidelines (clinic-published); ongoing leucovorin RCT enrollment"),
    ("Rossignol", "rossignolmedicalcenter.com / podcasts", "2025",
     "HBOT response biomarker pre-treatment work; clinic protocols"),
    ("Naviaux", "naviauxlab.ucsd.edu", "2026",
     "'3-hit metabolic signaling model' — pre-ASD biomarker framework via metabolomics; Mitochondrion 2026"),
    ("Hazan", "Twitter @SabinehazanMD", "2024-2025",
     "Bifidobacterium infantis depletion in autism; FMT/MTT clinical case reports"),
    ("Exley", "drchristopherexley.substack.com", "2024-2025",
     "'Aluminium and Autism: The Evidence' (July 2025); 'Aluminium and Autism: Unequivocal'; "
     "'Aluminium, Autism and Epilepsy'; 'Tylenol, Autism and All That' — sustained 2025 Substack series"),
    ("Lyons-Weiler", "popularrationalism.substack.com", "2025",
     "'CDC's Autism Reversal' (Nov 2025) — reports CDC website removed/modified 'Vaccines Do Not Cause "
     "Autism' language Nov 19 2025; 'Autism Health Summit' April 2025 organized; rebuttals to Atlantic, Leana Wen"),
    ("Hooker", "CHD.TV / childrenshealthdefense.org", "2024-2025",
     "Continued Thompson document advocacy; CHD-aligned commentary; no new peer-reviewed primary research"),
    ("Rogers", "tobyrogers.substack.com", "2025",
     "'Mapping the entire field of autism causation studies in one article' (May 2025) — comprehensive lit review; "
     "September 2025 Senate hearing testimony; ongoing Brownstone Institute contributions on regulatory capture"),
    ("Seneff", "Lectures + Substack", "2024",
     "PIN1 deficiency hypothesis (J Neurochem 2024); ongoing glyphosate/sulfur metabolism work"),
    ("Tomljenovic", "Independent", "2024-2025",
     "Less recent activity since UBC departure"),
    ("Blaylock", "russellblaylockmd.com / podcast", "2024-2025",
     "Continued immunoexcitotoxicity framework discussion; Substack-adjacent commentary"),
    ("Kory", "flccc.net", "2024-2025",
     "FLCCC PANS/autism protocol development; clinical case reports via Leading Edge Clinic"),
    ("Kulldorff", "kulldorff.com / Brownstone", "2024-2025",
     "Vaccine surveillance methodology critique; biostatistical writing on adverse event detection signals"),
    ("McCullough", "mcculloughfnd.org / Zenodo", "October 2025",
     "MCCULLOUGH FOUNDATION REPORT: 'Determinants of Autism Spectrum Disorder' — Zenodo "
     "DOI 10.5281/zenodo.17451259 (Oct 2025). Reviews 136 vaccine-autism studies (107 inferred link, "
     "29 neutral). NOT peer-reviewed (Zenodo = open repository, preprint-tier). Planned regression-"
     "after-vaccination case-series ongoing."),
    ("Walsh", "walshinstitute.org / Nutrient Power book", "2024-2025",
     "Walsh Research Institute clinical training programs; biotyping framework continued"),
    ("Siri", "@AaronSiriSG / sirilabel.com", "2024-2025",
     "Continued FOIA litigation; multiple wins surfacing CDC/FDA documents; ICAN-driven document releases"),
    ("Holland", "CHD.TV", "2024-2025",
     "Vaccine-injury legal/policy commentary; CHD presidency activities"),
    ("Kennedy", "HHS official channels + X", "2025",
     "HHS Secretary actions: ACIP autism review (Sep 2025); MAHA initiative; Hep B birth-dose policy review; "
     "federal autism initiative announcements"),
    ("Pfeiffer", "@marybethpf / marybethpfeiffer.com", "2024-2025",
     "Continued investigative reporting on chronic illness epidemic, Lyme, autism; FOIA-document journalism"),
    ("Tommey", "CHD.TV / VAXXED III", "2024-2025",
     "VAXXED III documentary release; ongoing parent-experience aggregation; Senate testimony Sep 2025"),
    ("Goh", "Cortica / corticacare.com", "2024-2025",
     "Cortica clinical protocols; whole-child autism medicine framework; integrated clinical model"),
    ("Geschwind", "geschwindlab.dgsom.ucla.edu", "2025-2026",
     "Multiple Nature papers 2025-2026 on developmental convergence/divergence in stem cell ASD models, "
     "polygenic profiles by age at diagnosis. Tier-1 mainstream genomics."),
    ("Adams", "autismnrc.org / ASU", "2024-2025",
     "PRO-S validation for PANS; p-cresol elevation review (2025); indoxyl sulfate review; "
     "Candida bimodal distribution; therapeutic diet effectiveness ratings (2023)"),
    ("Theoharides", "mastcellmaster.com", "2024",
     "Hippocampal molecular profiling in autism (Mol Psychiatry 2024); ochratoxin-induced mast cell IL-1β/IL-18/CXCL8 release"),
    ("Hollander", "Einstein/Montefiore lab", "2023-2024",
     "Oxytocin in neurodevelopmental disorders review (Pharmacol Ther 2024); cognitive inflexibility + immunome biomarkers; "
     "balovaptan/V1a clinical trial predictors"),
]

# === Atlas-ingested key papers ===
key_papers = [
    ("Frye", 2018, "30097774", "Mol Psychiatry", "Folinic acid improves verbal communication in autism (RCT)", "SRC-000001", 0.5225, "Atlas calibration anchor RCT"),
    ("Frye", 2017, "28250023", "(varies)", "Cerebral folate receptor autoantibodies in autism", "SRC-000002", 0.5, "Foundational FOLR1 work"),
    ("Frye", 2021, "34834493", "Pediatr Rev", "Treatment of folate metabolism abnormalities in ASD", "SRC-000011", 0.5, "Atlas-ingested"),
    ("Rossignol", 2009, "(multiple)", "Med Hypotheses", "HBOT for autism — early protocol papers", "SRC related", 0.4, "Multiple HBOT papers"),
    ("Naviaux", 2017, "28503133", "Ann Clin Transl Neurol", "SAT1: low-dose suramin in autism", "(not yet ingested)", 0.4, "N=10 open-label, mechanistically novel"),
    ("Naviaux", 2024, "38729981", "Communications Biology", "Metabolic network analysis of pre-ASD newborns and 5-yr-olds", "(not yet ingested)", 0.45, "Pre-ASD biomarker work"),
    ("Naviaux", 2026, "41902612", "Autism Research", "3-Hit Metabolic Signaling Model for ASD", "(not yet ingested)", 0.5, "Major framework update"),
    ("Theoharides", 2024, "38355786", "Mol Psychiatry", "Molecular profiling of hippocampus in autism", "(not yet ingested)", 0.5, "Mainstream Mol Psychiatry"),
    ("Theoharides", 2024, "38301823", "Toxicology", "Ochratoxin A → mast cell IL-1β/IL-18/CXCL8 release", "(not yet ingested)", 0.45, "Mycotoxin → MCAS link"),
    ("Hazan", 2024, "38715916", "J Med Cases", "Gut microbiome composition + clinical symptoms post-MTT", "(not yet ingested)", 0.35, "Clinical case report"),
    ("Adams", 2017, "29076465", "Sci Rep", "MTT for autism (open-label, 2-year)", "(not yet ingested)", 0.5, "Foundational MTT"),
    ("Adams", 2025, "40863715", "Pediatr Rep", "PANS-Recurrent Outcome Survey (PRO-S) validation", "(not yet ingested)", 0.45, "Direct atlas Session 3.5C feed"),
    ("Adams", 2025, "40003979", "Int J Mol Sci", "Para-cresol elevation in autism — review", "(not yet ingested)", 0.4, "Direct biomarker layer feed"),
    ("Adams", 2024, "39684683", "Int J Mol Sci", "Indoxyl sulfate and ASD review", "(not yet ingested)", 0.4, "Biomarker review"),
    ("Adams", 2024, "41907709", "Gut Microbes Reports", "Bimodal distribution of intestinal Candida in autism", "(not yet ingested)", 0.4, "Subpopulation finding"),
    ("Exley", 2018, "29161691", "J Trace Elem Med Biol", "Aluminum in brain tissue in autism", "(not yet ingested)", 0.4, "Mochizuki-Exley foundational"),
    ("Hooker", 2014, "24995277", "Biomed Res Int", "Methodological issues thimerosal-safety research", "SRC-001434", 0.15, "Atlas-ingested at low strength"),
    ("Geschwind", 2018, "30382198", "Science", "Transcriptome-wide isoform-level dysregulation in ASD", "(connects via gene layer)", 0.6, "PsychENCODE foundational"),
    ("Geschwind", 2026, "41611887", "Nature", "Developmental convergence/divergence in stem cell ASD models", "(not yet ingested)", 0.65, "Tier-1 Nature 2026"),
    ("Geschwind", 2025, "41034588", "Nature", "Polygenic and developmental profiles of autism by age at diagnosis", "(not yet ingested)", 0.65, "Tier-1 Nature 2025"),
    ("Hollander", 2024, "39455012", "Pharmacol Ther", "Oxytocin in neurodevelopmental disorders: ASD and Prader-Willi", "(not yet ingested)", 0.4, "Direct Session 3.5G feed"),
    ("Tomljenovic", 2011, "22099159", "J Inorg Biochem", "Aluminum vaccine adjuvants and autism (Hill criteria)", "SRC-001431", 0.30, "Atlas-ingested"),
    ("Blaylock", 2008, "19043938", "Altern Ther Health Med", "Immunoexcitotoxicity central mechanism in ASD", "SRC-001444", 0.10, "Atlas-ingested at low strength"),
    ("Seneff", 2024, "38808598", "J Neurochem", "Is autism a PIN1 deficiency syndrome?", "(not yet ingested)", 0.2, "Novel framework 2024"),
    ("Adams", 2023, "37888059", "J Pers Med", "Effectiveness ratings of 13 therapeutic diets for ASD", "(not yet ingested)", 0.4, "Diet ranking"),
    ("Goh", 2025, "39092588", "Genetics in Medicine", "Penetrance estimates of CNVs - systematic review", "(not yet ingested)", 0.45, "CNV penetrance synthesis"),
    ("Frye", 2026, "41751904", "Int J Mol Sci", "Gut-Brain Inflammation and Disrupted Homeostasis", "(not yet ingested)", 0.5, "Recent gut-brain work"),
    ("Frye", 2025, "41010044", "Genes", "De Novo Variants Predominate in ASD", "(not yet ingested)", 0.5, "Genetic architecture"),
    ("Frye", 2025, "41010010", "Genes", "Transcriptomic Signatures of Mitochondrial Dysfunction in ASD", "(not yet ingested)", 0.5, "Mito transcriptomics"),
    ("Frye", 2025, "40943215", "Int J Mol Sci", "Transgenerational Effects + Heritability of FOLR1 Autoantibodies", "(not yet ingested)", 0.5, "FOLR1 inheritance — major"),
    ("Rossignol", 2024, "38703861", "Neurobiol Dis", "Biomarkers of mitochondrial dysfunction in autism", "(not yet ingested)", 0.5, "Biomarker review"),
    ("Rossignol", 2024, "38248763", "J Pers Med", "Folate Receptor Alpha Autoantibody = leucovorin response biomarker", "(not yet ingested)", 0.5, "Leucovorin stratification"),
    ("Rossignol", 2025, "40481745", "Expert Rev Mol Diagn", "Early biomarker for ASD unveiled", "(not yet ingested)", 0.5, "Recent review"),
]


# === Signal vs Noise ===
signal_noise = [
    ("Frye", 9, "Posts citing his own RCTs; FOLR1 + leucovorin discussion; cerebral folate phenotype refinement; PANS treatment protocols", "Almost no noise — research-focused", "Default: read everything"),
    ("Rossignol", 7, "Posts on mito biomarker stratification; HBOT response prediction; folate receptor antibody biomarker", "Limited Twitter; minimal noise", "Read all (low volume)"),
    ("Naviaux", 8, "Suramin trial updates; CDR theory; metabolomics findings; 3-hit model", "Rare commentary", "Read all (low volume)"),
    ("Theoharides", 8, "Mast cell biology in autism; flavonoid mechanism; mycotoxin links", "Some commercial luteolin promotion", "Filter for non-commercial science"),
    ("Hazan", 7, "Bifidobacterium findings; FMT protocols; primary microbiome work", "Some commercial-adjacent", "Filter: keep posts citing primary papers"),
    ("Adams", 9, "PANS validation, p-cresol, Candida bimodal, indoxyl sulfate — almost everything is signal", "Limited X presence", "Subscribe via ASU page rather than Twitter"),
    ("Exley", 8, "Aluminum biochemistry findings; Mochizuki-Exley brain Al data; sustained Substack 2025", "Career-loss content; advocacy fundraising", "Filter: keep biochemistry; skip personal-circumstance"),
    ("Lyons-Weiler", 6, "Peer-reviewed paper announcements; specific findings; CDC website language change reporting", "Frequent advocacy retweets; engagement-bait", "Filter heavily: keep posts linking to PMID"),
    ("Hooker", 5, "Thompson document references; FOIA-document announcements", "2014 retraction reduces methodology trust", "Trust on FOIA; not on epidemiology"),
    ("Rogers", 7, "Substack long-form regulatory analysis; Senate testimony references; May 2025 mapping article", "Less noise; limited biomedical signal", "Read substack long-form; ignore brevity"),
    ("Seneff", 5, "Specific glyphosate-methionine biochemistry; sulfur metabolism observations; PIN1 hypothesis 2024", "Apocalyptic predictions; over-broad claims", "Filter heavily: keep narrow biochemistry"),
    ("Geschwind", 9, "Lab paper announcements; PsychENCODE; SFARI Tier 1 findings; multiple Nature papers 2025-2026", "Almost no Twitter content", "Track via PubMed alerts on his lab"),
    ("Hollander", 7, "Trial recruitment; clinical trial outcomes; oxytocin/vasopressin", "Limited Twitter activity", "Track via clinicaltrials.gov + lab page"),
    ("Siri", 8, "FOIA-released document announcements; legal filings producing primary documents", "Some advocacy framing", "Highest priority for FOIA signals"),
    ("Holland", 6, "Legal/policy framing of vaccine injury; CHD documentation", "CHD-aligned advocacy", "Use for legal/policy; not biomedical"),
    ("Kennedy", 5, "HHS policy actions; ACIP review announcements; federal autism initiative", "Frequent rhetorical inflation of biomedical claims", "Trust for: policy news. Not for: biomedical specifics"),
    ("Pfeiffer", 7, "Investigative-journalism long-form; FOIA document surfacing", "Limited noise (rigorous reporter)", "High trust for sourced reporting"),
    ("Tommey", 4, "Parent-experience aggregation; documentary content", "Heavy advocacy framing", "Use sparingly; subjective-response only"),
    ("McCullough", 6, "McCullough Foundation autism case series; 2025 Zenodo Determinants of ASD report", "COVID-era engagement still dominates", "Watch for autism Foundation outputs"),
    ("Walsh", 6, "Walsh biotyping framework; clinical-experience aggregation", "Limited Twitter", "Use Walsh Research Institute publications"),
    ("Tomljenovic", 6, "Aluminum adjuvant primary research findings", "Career-loss content; less recent", "Track historical primary research"),
    ("Blaylock", 5, "Immunoexcitotoxicity framework references", "Heavy advocacy/alt-med engagement", "Use framework references; filter advocacy"),
    ("Goh", 7, "Cortica clinical protocols; whole-child autism medicine; CNV penetrance work", "Cortica-commercial alignment", "Track Cortica protocols; filter commercial framing"),
    ("Kory", 5, "FLCCC PANS/autism extension protocols", "Heavy COVID-era residual content", "Filter heavily for autism-specific"),
    ("Kulldorff", 8, "Vaccine-surveillance methodology critique; biostatistics rigor", "Some COVID-era policy positioning", "Highest trust for biostat methodology critique"),
]

# Build workbook
wb = Workbook()

def make_header_row(ws, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

def style_body(ws, n_rows, n_cols):
    for row_idx in range(2, n_rows+2):
        for col_idx in range(1, n_cols+1):
            cell = ws.cell(row_idx, col_idx)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER

# --- Sheet 1: Researchers ---
sh1 = wb.active
sh1.title = "Researchers"
make_header_row(sh1, [
    "#", "Name", "Credentials", "Affiliation", "Twitter/X handle",
    "Other primary platform", "Parent of ASD child?",
    "Research focus", "Key contributions",
    "Atlas connections", "Tier",
    "Signal density (1-10)", "Bias considerations", "Reliability",
    "Notes / recent activity"
])
for row_idx, r in enumerate(researchers, 2):
    for col_idx, val in enumerate(r, 1):
        sh1.cell(row_idx, col_idx, val)
style_body(sh1, len(researchers), 15)
for row_idx, r in enumerate(researchers, 2):
    tier = r[10]
    if tier in TIER_FILLS:
        for col_idx in range(1, len(r)+1):
            sh1.cell(row_idx, col_idx).fill = TIER_FILLS[tier]
widths1 = [4, 22, 12, 35, 22, 28, 6, 35, 50, 35, 28, 8, 40, 12, 50]
for i, w in enumerate(widths1, 1):
    sh1.column_dimensions[get_column_letter(i)].width = w
sh1.freeze_panes = "A2"
sh1.row_dimensions[1].height = 36
for row_idx in range(2, len(researchers)+2):
    sh1.row_dimensions[row_idx].height = 110

# --- Sheet 2: Recent Publications (PubMed) ---
sh2 = wb.create_sheet("Recent Publications")
make_header_row(sh2, ["Researcher", "Year", "PMID", "Journal (truncated)",
                       "Title (truncated)", "First Author"])
for row_idx, r in enumerate(recent_pubs_table, 2):
    for col_idx, val in enumerate(r, 1):
        sh2.cell(row_idx, col_idx, val)
style_body(sh2, len(recent_pubs_table), 6)
widths2 = [18, 8, 14, 38, 70, 18]
for i, w in enumerate(widths2, 1):
    sh2.column_dimensions[get_column_letter(i)].width = w
sh2.freeze_panes = "A2"
for row_idx in range(2, len(recent_pubs_table)+2):
    sh2.row_dimensions[row_idx].height = 30

# --- Sheet 3: Recent Substack/Public Findings ---
sh3 = wb.create_sheet("Recent Substack & Public")
make_header_row(sh3, ["Researcher", "Primary Platform", "Date Range",
                       "Recent activity / key findings"])
for row_idx, r in enumerate(substack_findings, 2):
    for col_idx, val in enumerate(r, 1):
        sh3.cell(row_idx, col_idx, val)
style_body(sh3, len(substack_findings), 4)
widths3 = [18, 30, 14, 90]
for i, w in enumerate(widths3, 1):
    sh3.column_dimensions[get_column_letter(i)].width = w
sh3.freeze_panes = "A2"
for row_idx in range(2, len(substack_findings)+2):
    sh3.row_dimensions[row_idx].height = 60

# --- Sheet 4: Key Atlas-Ingested Papers ---
sh4 = wb.create_sheet("Key Papers")
make_header_row(sh4, ["Researcher", "Year", "PMID", "Journal",
                       "Title", "Atlas SRC if ingested",
                       "Strength score", "Notes"])
for row_idx, r in enumerate(key_papers, 2):
    for col_idx, val in enumerate(r, 1):
        sh4.cell(row_idx, col_idx, val)
style_body(sh4, len(key_papers), 8)
widths4 = [18, 8, 14, 22, 50, 22, 14, 35]
for i, w in enumerate(widths4, 1):
    sh4.column_dimensions[get_column_letter(i)].width = w
sh4.freeze_panes = "A2"

# --- Sheet 5: Signal vs Noise Guide ---
sh5 = wb.create_sheet("Signal vs Noise Guide")
make_header_row(sh5, ["Researcher", "Signal density (1-10)",
                       "What to weight as SIGNAL",
                       "What to filter as NOISE", "Practical filter rule"])
for row_idx, r in enumerate(signal_noise, 2):
    for col_idx, val in enumerate(r, 1):
        sh5.cell(row_idx, col_idx, val)
style_body(sh5, len(signal_noise), 5)
widths5 = [18, 8, 50, 45, 40]
for i, w in enumerate(widths5, 1):
    sh5.column_dimensions[get_column_letter(i)].width = w
sh5.freeze_panes = "A2"
sh5.row_dimensions[1].height = 36
for row_idx in range(2, len(signal_noise)+2):
    sh5.row_dimensions[row_idx].height = 60

# --- Sheet 6: Methodology + Atlas Linkage ---
sh6 = wb.create_sheet("Methodology + Atlas Linkage")
sh6.column_dimensions['A'].width = 110
methodology = [
    "Autism Researchers Tracker — Methodology Notes",
    "",
    f"Generated: {datetime.now().strftime('%Y-%m-%d')}",
    "v2: enriched with PubMed recent publications (2023-2026) + Substack/public-source recent activity",
    "",
    "PURPOSE",
    "Track 25 autism researchers + parent-advocates whose primary research output the atlas should ingest.",
    "Per CLAUDE.md mainstream-skeptical principle, the list spans mainstream academic, functional medicine,",
    "contested-evidence, and parent-researcher tiers. Their PEER-REVIEWED OUTPUT is the signal; Twitter is",
    "downstream commentary that points toward what they're working on.",
    "",
    "KEY MAJOR DEVELOPMENTS CAPTURED IN THIS VERSION",
    "1. CDC website language change Nov 19 2025 — 'Vaccines Do Not Cause Autism' modified per Lyons-Weiler reporting",
    "2. McCullough Foundation Determinants of ASD report October 2025 (Zenodo, preprint-tier, 136 studies reviewed)",
    "3. Naviaux 2026 '3-Hit Metabolic Signaling Model' for ASD (Mitochondrion + Autism Research)",
    "4. Geschwind multiple Nature papers 2025-2026 on developmental ASD genomics",
    "5. Theoharides 2024 ochratoxin → mast cell IL-1β/IL-18/CXCL8 release (mycotoxin-MCAS-autism link)",
    "6. Adams 2024-2025 cluster: PANS validation, p-cresol, indoxyl sulfate, Candida bimodal — directly feeds Session 3.5A biomarker layer",
    "7. Frye 2026 transgenerational FOLR1 autoantibody heritability paper",
    "8. Hollander 2024 oxytocin/vasopressin in neurodevelopmental disorders (Pharmacol Ther)",
    "9. Exley sustained 2024-2025 Substack series on aluminum-autism (drchristopherexley.substack.com)",
    "10. Rogers May 2025 'Mapping the entire field of autism causation studies' Substack synthesis",
    "",
    "SIGNAL DENSITY SCALE (1-10)",
    "9-10: Almost all public output is research-grounded (Frye, Adams, Geschwind)",
    "7-8: Mostly research-grounded with occasional commentary (Naviaux, Theoharides, Kulldorff, Siri)",
    "5-6: Mixed; signal embedded in commentary (Lyons-Weiler, Walsh, Holland)",
    "3-4: Heavy advocacy/engagement framing (Tommey)",
    "",
    "RELIABILITY TIERS (per CLAUDE.md source-quality framework)",
    "Very High: Mainstream academic with clean track record (Geschwind, Adams, Kulldorff)",
    "High: Functional medicine with strong primary research (Frye, Naviaux, Theoharides, Rossignol, Siri)",
    "Medium-High: Some mixed elements but substantive primary work (Hazan, Exley, Pfeiffer, Goh, Walsh)",
    "Medium: Mixed; needs filtering (Lyons-Weiler, Rogers, Holland, Seneff, McCullough, Tomljenovic, Blaylock, Kory, Kennedy, Tommey)",
    "Medium-Low: Strong contested-evidence advocacy alignment (Hooker)",
    "",
    "HOW TO USE THIS TRACKER",
    "1. For each researcher, set up a PubMed alert on their name + 'autism' or 'ASD'",
    "2. Set up Substack subscriptions for those whose primary platform is Substack",
    "3. When they publish a new paper, the paper itself goes into the atlas via run_ingest.py pmid",
    "4. Per CLAUDE.md, the atlas weights based on study design + source type, NOT author popularity",
    "5. Researchers with high signal density + high reliability are highest priority for ingestion",
    "",
    "PARENT-OF-AUTISTIC-CHILD STATUS",
    "Confirmed parents: Adams (autistic adult daughter), Hooker (vaccine-injured son), Tommey (autistic son),",
    "Holland (autistic child)",
    "RFK Jr.: extended family advocacy (no direct parental relationship to my knowledge)",
    "",
    "ATLAS CALIBRATION",
    "INT-0001 Leucovorin = 83.20 CSRS (calibration anchor)",
    "Calibration anchor researcher: Frye (Frye/Slattery FOLR1 work)",
    "Atlas current state: 73 hypotheses, 34 mechanisms, 101 interventions, 7 phenotypes, 1428 sources, 1564 genes (100% wired)",
    "",
    "TWITTER/X CONTENT LIMITATIONS",
    "X content is JS-rendered behind authentication walls; programmatic extraction infeasible without auth.",
    "WORKAROUND: This tracker captures their PUBLISHED primary research (PubMed) + Substack/blog recent",
    "activity (publicly accessible). For real-time Twitter signal, follow accounts directly in your X app.",
    "",
    "NEXT STEPS PER ATLAS ROADMAP (CLAUDE.md Session 3.5+)",
    "- Session 3.5A: ingest functional medicine biomarkers — Adams 2024-2025 papers feed directly",
    "- Session 3.5B: ingest functional medicine interventions — Frye/Rossignol/Theoharides papers",
    "- Session 3.5C: PANS/MCAS hypothesis families — Theoharides (MCAS) + Adams PRO-S (PANS) primary",
    "- Session 3.5D: Walsh/Pfeiffer phenotype refinement — Walsh primary",
    "- Session 3.5E: named-protocol entities — Frye, Walsh, Bock, Klinghardt, MAPS",
    "- Session 3.5F: drug repurposing — Naviaux suramin, etc.",
    "- Session 3.5G: peptide layer — Hollander oxytocin work primary",
]
for row_idx, line in enumerate(methodology, 1):
    cell = sh6.cell(row_idx, 1, line)
    if row_idx == 1:
        cell.font = Font(name='Arial', size=14, bold=True, color='1F4E79')
    elif line in ("PURPOSE", "KEY MAJOR DEVELOPMENTS CAPTURED IN THIS VERSION", "SIGNAL DENSITY SCALE (1-10)",
                  "RELIABILITY TIERS (per CLAUDE.md source-quality framework)", "HOW TO USE THIS TRACKER",
                  "PARENT-OF-AUTISTIC-CHILD STATUS", "ATLAS CALIBRATION",
                  "TWITTER/X CONTENT LIMITATIONS", "NEXT STEPS PER ATLAS ROADMAP (CLAUDE.md Session 3.5+)"):
        cell.font = Font(name='Arial', size=11, bold=True, color='1F4E79')
    else:
        cell.font = Font(name='Arial', size=10)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheet 1 'Researchers': {len(researchers)} rows")
print(f"Sheet 2 'Recent Publications': {len(recent_pubs_table)} rows (PubMed-pulled 2023-2026)")
print(f"Sheet 3 'Recent Substack & Public': {len(substack_findings)} rows")
print(f"Sheet 4 'Key Papers': {len(key_papers)} rows")
print(f"Sheet 5 'Signal vs Noise Guide': {len(signal_noise)} rows")
print(f"Sheet 6 'Methodology + Atlas Linkage': overview")
